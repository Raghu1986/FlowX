from openpyxl import load_workbook
from io import BytesIO

def stream_excel_records(excel_bytes):
    """Yield rows as dicts with lowercase headers (case-insensitive matching)."""
    wb = load_workbook(excel_bytes, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
